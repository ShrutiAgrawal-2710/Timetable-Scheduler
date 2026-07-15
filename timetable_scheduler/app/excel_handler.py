import io
import pandas as pd
from datetime import datetime
from typing import List, Dict, Tuple, Any
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class ExcelValidationError(Exception):
    def __init__(self, errors: List[Dict[str, Any]]):
        self.errors = errors
        super().__init__(f"Excel validation failed with {len(errors)} errors")

def generate_template() -> io.BytesIO:
    """
    Generates a pre-formatted empty Excel template with instructions and sample rows.
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    
    # Styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=11)
    title_font = Font(name="Segoe UI", size=14, bold=True, color="1F4E78")
    instruction_font = Font(name="Segoe UI", size=10, italic=True, color="595959")
    
    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    
    # 1. General Settings Sheet
    ws_settings = wb.create_sheet(title="General_Settings")
    ws_settings.views.sheetView[0].showGridLines = True
    
    ws_settings.append(["Timetable Scheduler - General Settings"])
    ws_settings.cell(row=1, column=1).font = title_font
    ws_settings.append(["Define the planning horizon and global configurations below."])
    ws_settings.cell(row=2, column=1).font = instruction_font
    ws_settings.append([]) # empty row
    
    headers_settings = ["Parameter", "Value", "Description"]
    ws_settings.append(headers_settings)
    
    settings_data = [
        ["Start Date", "2026-08-01", "Start date of the schedule planning period (YYYY-MM-DD)"],
        ["End Date", "2026-08-14", "End date of the schedule planning period (YYYY-MM-DD)"],
        ["Default Shift Hours", 8.0, "Standard work hours for a single day shift assignment"]
    ]
    for row in settings_data:
        ws_settings.append(row)
        
    # Apply style to General Settings
    for col_idx in range(1, len(headers_settings) + 1):
        cell = ws_settings.cell(row=4, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for r_idx in range(5, 5 + len(settings_data)):
        for col_idx in range(1, len(headers_settings) + 1):
            cell = ws_settings.cell(row=r_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            if col_idx == 1:
                cell.font = Font(name="Segoe UI", size=11, bold=True)
                
    # 2. Resources Sheet
    ws_res = wb.create_sheet(title="Resources")
    ws_res.views.sheetView[0].showGridLines = True
    ws_res.append(["Resources / Persons Master List"])
    ws_res.cell(row=1, column=1).font = title_font
    ws_res.append(["List all individuals to be scheduled and their planning period hour limits."])
    ws_res.cell(row=2, column=1).font = instruction_font
    ws_res.append([])
    
    headers_res = ["Resource Name", "Min Hours", "Max Hours"]
    ws_res.append(headers_res)
    sample_res = [
        ["Dr. Alice Smith", 40.0, 80.0],
        ["Dr. Bob Jones", 30.0, 60.0],
        ["Dr. Charlie Brown", 40.0, 80.0]
    ]
    for row in sample_res:
        ws_res.append(row)
        
    for col_idx in range(1, len(headers_res) + 1):
        cell = ws_res.cell(row=4, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        
    for r_idx in range(5, 5 + len(sample_res)):
        for col_idx in range(1, len(headers_res) + 1):
            cell = ws_res.cell(row=r_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            
    # 3. Departments Sheet
    ws_dept = wb.create_sheet(title="Departments")
    ws_dept.views.sheetView[0].showGridLines = True
    ws_dept.append(["Departments / Tasks Master List"])
    ws_dept.cell(row=1, column=1).font = title_font
    ws_dept.append(["List departments/courses needing scheduling and their total required hours."])
    ws_dept.cell(row=2, column=1).font = instruction_font
    ws_dept.append([])
    
    headers_dept = ["Department Name", "Min Hours Needed", "Max Hours Needed"]
    ws_dept.append(headers_dept)
    sample_dept = [
        ["Cardiology", 60.0, 120.0],
        ["Pediatrics", 40.0, 80.0],
        ["Emergency Room", 80.0, 160.0]
    ]
    for row in sample_dept:
        ws_dept.append(row)
        
    for col_idx in range(1, len(headers_dept) + 1):
        cell = ws_dept.cell(row=4, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        
    for r_idx in range(5, 5 + len(sample_dept)):
        for col_idx in range(1, len(headers_dept) + 1):
            cell = ws_dept.cell(row=r_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            
    # 4. Resource Dept Mapping Sheet
    ws_map = wb.create_sheet(title="Resource_Dept_Mapping")
    ws_map.views.sheetView[0].showGridLines = True
    ws_map.append(["Resource Department Feasibility & Bounds"])
    ws_map.cell(row=1, column=1).font = title_font
    ws_map.append(["Define which resources can serve which departments, and day bounds."])
    ws_map.cell(row=2, column=1).font = instruction_font
    ws_map.append([])
    
    headers_map = ["Resource Name", "Department Name", "Is Feasible (Yes/No)", "Min Days", "Max Days"]
    ws_map.append(headers_map)
    sample_map = [
        ["Dr. Alice Smith", "Cardiology", "Yes", 2, 10],
        ["Dr. Alice Smith", "Emergency Room", "Yes", 0, 5],
        ["Dr. Bob Jones", "Pediatrics", "Yes", 2, 8],
        ["Dr. Bob Jones", "Emergency Room", "Yes", 1, 5],
        ["Dr. Charlie Brown", "Cardiology", "Yes", 0, 5],
        ["Dr. Charlie Brown", "Emergency Room", "Yes", 3, 10]
    ]
    for row in sample_map:
        ws_map.append(row)
        
    for col_idx in range(1, len(headers_map) + 1):
        cell = ws_map.cell(row=4, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        
    for r_idx in range(5, 5 + len(sample_map)):
        for col_idx in range(1, len(headers_map) + 1):
            cell = ws_map.cell(row=r_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            
    # 5. Absences Sheet
    ws_abs = wb.create_sheet(title="Absences")
    ws_abs.views.sheetView[0].showGridLines = True
    ws_abs.append(["Resource Absence Planning"])
    ws_abs.cell(row=1, column=1).font = title_font
    ws_abs.append(["Define specific dates when a resource is absent. Saturdays and Sundays are absent by default."])
    ws_abs.cell(row=2, column=1).font = instruction_font
    ws_abs.append([])
    
    headers_abs = ["Resource Name", "Absence Date", "Reason"]
    ws_abs.append(headers_abs)
    sample_abs = [
        ["Dr. Alice Smith", "2026-08-05", "Medical Conference"],
        ["Dr. Bob Jones", "2026-08-10", "Personal Leave"]
    ]
    for row in sample_abs:
        ws_abs.append(row)
        
    for col_idx in range(1, len(headers_abs) + 1):
        cell = ws_abs.cell(row=4, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        
    for r_idx in range(5, 5 + len(sample_abs)):
        for col_idx in range(1, len(headers_abs) + 1):
            cell = ws_abs.cell(row=r_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border

    # Auto-adjust column widths for all sheets
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            # Skip the first two rows (title/instructions) for width calculation to avoid overly wide columns
            for cell in col[3:]:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 15)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def parse_and_validate_excel(file_bytes: bytes) -> Dict[str, Any]:
    """
    Parses and validates the timetable scheduling Excel sheet.
    Returns parsed dictionary of structured data or raises ExcelValidationError.
    """
    errors = []
    
    try:
        xls = pd.ExcelFile(io.BytesIO(file_bytes))
    except Exception as e:
        raise ExcelValidationError([{"sheet": "File", "row": "N/A", "column": "N/A", "message": f"Could not read file. Is it a valid Excel workbook? Details: {str(e)}"}])
        
    required_sheets = ["General_Settings", "Resources", "Departments", "Resource_Dept_Mapping", "Absences"]
    for sheet in required_sheets:
        if sheet not in xls.sheet_names:
            errors.append({"sheet": sheet, "row": "N/A", "column": "N/A", "message": f"Missing required tab: '{sheet}'"})
            
    if errors:
        raise ExcelValidationError(errors)
        
    # 1. Parse General Settings
    df_settings = pd.read_excel(xls, sheet_name="General_Settings", skiprows=3)
    expected_settings_cols = ["Parameter", "Value"]
    for col in expected_settings_cols:
        if col not in df_settings.columns:
            errors.append({"sheet": "General_Settings", "row": "Header", "column": col, "message": f"Missing column '{col}' in General_Settings"})
            
    start_date = None
    end_date = None
    default_shift_hours = 8.0
    
    if not errors:
        for idx, row in df_settings.iterrows():
            param = str(row["Parameter"]).strip()
            val = row["Value"]
            row_num = idx + 5 # 4 rows skipped/header
            
            if param == "Start Date":
                try:
                    if isinstance(val, (datetime, pd.Timestamp)):
                        start_date = val.date()
                    else:
                        start_date = datetime.strptime(str(val).strip(), "%Y-%m-%d").date()
                except Exception:
                    errors.append({"sheet": "General_Settings", "row": row_num, "column": "Value", "message": f"Invalid Start Date format '{val}'. Expected YYYY-MM-DD."})
            elif param == "End Date":
                try:
                    if isinstance(val, (datetime, pd.Timestamp)):
                        end_date = val.date()
                    else:
                        end_date = datetime.strptime(str(val).strip(), "%Y-%m-%d").date()
                except Exception:
                    errors.append({"sheet": "General_Settings", "row": row_num, "column": "Value", "message": f"Invalid End Date format '{val}'. Expected YYYY-MM-DD."})
            elif param == "Default Shift Hours":
                try:
                    default_shift_hours = float(val)
                    if default_shift_hours <= 0:
                        errors.append({"sheet": "General_Settings", "row": row_num, "column": "Value", "message": "Default Shift Hours must be greater than zero."})
                except Exception:
                    errors.append({"sheet": "General_Settings", "row": row_num, "column": "Value", "message": f"Default Shift Hours must be a number (got '{val}')."})
        
        if start_date and end_date and start_date > end_date:
            errors.append({"sheet": "General_Settings", "row": "N/A", "column": "Value", "message": "Start Date cannot be after End Date."})

    # 2. Parse Resources
    df_res = pd.read_excel(xls, sheet_name="Resources", skiprows=3)
    res_cols = ["Resource Name", "Min Hours", "Max Hours"]
    for col in res_cols:
        if col not in df_res.columns:
            errors.append({"sheet": "Resources", "row": "Header", "column": col, "message": f"Missing column '{col}' in Resources"})
            
    resources = []
    resource_names = set()
    if not errors:
        for idx, row in df_res.iterrows():
            row_num = idx + 5
            name = str(row["Resource Name"]).strip()
            min_h = row["Min Hours"]
            max_h = row["Max Hours"]
            
            if pd.isna(row["Resource Name"]) or name == "":
                errors.append({"sheet": "Resources", "row": row_num, "column": "Resource Name", "message": "Resource Name cannot be empty."})
                continue
                
            if name in resource_names:
                errors.append({"sheet": "Resources", "row": row_num, "column": "Resource Name", "message": f"Duplicate resource name '{name}'."})
            resource_names.add(name)
            
            try:
                min_h = float(min_h) if not pd.isna(min_h) else 0.0
                max_h = float(max_h) if not pd.isna(max_h) else 240.0
                if min_h < 0 or max_h < 0:
                    errors.append({"sheet": "Resources", "row": row_num, "column": "Hours", "message": "Hours cannot be negative."})
                if min_h > max_h:
                    errors.append({"sheet": "Resources", "row": row_num, "column": "Min Hours", "message": "Min Hours cannot exceed Max Hours."})
            except Exception:
                errors.append({"sheet": "Resources", "row": row_num, "column": "Hours", "message": "Hours must be numbers."})
                
            resources.append({"name": name, "min_hours": min_h, "max_hours": max_h})

    # 3. Parse Departments
    df_dept = pd.read_excel(xls, sheet_name="Departments", skiprows=3)
    dept_cols = ["Department Name", "Min Hours Needed", "Max Hours Needed"]
    for col in dept_cols:
        if col not in df_dept.columns:
            errors.append({"sheet": "Departments", "row": "Header", "column": col, "message": f"Missing column '{col}' in Departments"})
            
    departments = []
    department_names = set()
    if not errors:
        for idx, row in df_dept.iterrows():
            row_num = idx + 5
            name = str(row["Department Name"]).strip()
            min_h = row["Min Hours Needed"]
            max_h = row["Max Hours Needed"]
            
            if pd.isna(row["Department Name"]) or name == "":
                errors.append({"sheet": "Departments", "row": row_num, "column": "Department Name", "message": "Department Name cannot be empty."})
                continue
                
            if name in department_names:
                errors.append({"sheet": "Departments", "row": row_num, "column": "Department Name", "message": f"Duplicate department name '{name}'."})
            department_names.add(name)
            
            try:
                min_h = float(min_h) if not pd.isna(min_h) else 0.0
                max_h = float(max_h) if not pd.isna(max_h) else 240.0
                if min_h < 0 or max_h < 0:
                    errors.append({"sheet": "Departments", "row": row_num, "column": "Hours", "message": "Hours cannot be negative."})
                if min_h > max_h:
                    errors.append({"sheet": "Departments", "row": row_num, "column": "Min Hours Needed", "message": "Min Hours Needed cannot exceed Max Hours Needed."})
            except Exception:
                errors.append({"sheet": "Departments", "row": row_num, "column": "Hours", "message": "Hours must be numbers."})
                
            departments.append({"name": name, "min_hours": min_h, "max_hours": max_h})

    # 4. Parse Resource Dept Mapping
    df_map = pd.read_excel(xls, sheet_name="Resource_Dept_Mapping", skiprows=3)
    map_cols = ["Resource Name", "Department Name", "Is Feasible (Yes/No)", "Min Days", "Max Days"]
    for col in map_cols:
        if col not in df_map.columns:
            errors.append({"sheet": "Resource_Dept_Mapping", "row": "Header", "column": col, "message": f"Missing column '{col}' in Resource_Dept_Mapping"})
            
    mappings = []
    if not errors:
        for idx, row in df_map.iterrows():
            row_num = idx + 5
            r_name = str(row["Resource Name"]).strip()
            d_name = str(row["Department Name"]).strip()
            feasible_str = str(row["Is Feasible (Yes/No)"]).strip().lower()
            min_d = row["Min Days"]
            max_d = row["Max Days"]
            
            if pd.isna(row["Resource Name"]) or r_name == "":
                errors.append({"sheet": "Resource_Dept_Mapping", "row": row_num, "column": "Resource Name", "message": "Resource Name cannot be empty."})
                continue
            if pd.isna(row["Department Name"]) or d_name == "":
                errors.append({"sheet": "Resource_Dept_Mapping", "row": row_num, "column": "Department Name", "message": "Department Name cannot be empty."})
                continue
                
            if r_name not in resource_names:
                errors.append({"sheet": "Resource_Dept_Mapping", "row": row_num, "column": "Resource Name", "message": f"Resource '{r_name}' is not defined in the Resources sheet."})
            if d_name not in department_names:
                errors.append({"sheet": "Resource_Dept_Mapping", "row": row_num, "column": "Department Name", "message": f"Department '{d_name}' is not defined in the Departments sheet."})
                
            is_feasible = True
            if feasible_str in ["no", "n", "false", "0"]:
                is_feasible = False
            elif feasible_str in ["yes", "y", "true", "1"]:
                is_feasible = True
            else:
                errors.append({"sheet": "Resource_Dept_Mapping", "row": row_num, "column": "Is Feasible (Yes/No)", "message": f"Invalid feasibility value '{row['Is Feasible (Yes/No)']}'; use Yes or No."})
                
            try:
                min_d = int(min_d) if not pd.isna(min_d) else 0
                max_d = int(max_d) if not pd.isna(max_d) else 31
                if min_d < 0 or max_d < 0:
                    errors.append({"sheet": "Resource_Dept_Mapping", "row": row_num, "column": "Days", "message": "Days cannot be negative."})
                if min_d > max_d:
                    errors.append({"sheet": "Resource_Dept_Mapping", "row": row_num, "column": "Min Days", "message": "Min Days cannot exceed Max Days."})
            except Exception:
                errors.append({"sheet": "Resource_Dept_Mapping", "row": row_num, "column": "Days", "message": "Days must be integers."})
                
            mappings.append({
                "resource_name": r_name,
                "dept_name": d_name,
                "is_feasible": is_feasible,
                "min_days": min_d,
                "max_days": max_d
            })

    # 5. Parse Absences
    df_abs = pd.read_excel(xls, sheet_name="Absences", skiprows=3)
    abs_cols = ["Resource Name", "Absence Date", "Reason"]
    for col in abs_cols:
        if col not in df_abs.columns:
            errors.append({"sheet": "Absences", "row": "Header", "column": col, "message": f"Missing column '{col}' in Absences"})
            
    absences = []
    if not errors:
        for idx, row in df_abs.iterrows():
            row_num = idx + 5
            r_name = str(row["Resource Name"]).strip()
            date_val = row["Absence Date"]
            reason = str(row["Reason"]).strip() if not pd.isna(row["Reason"]) else ""
            
            if pd.isna(row["Resource Name"]) or r_name == "":
                errors.append({"sheet": "Absences", "row": row_num, "column": "Resource Name", "message": "Resource Name cannot be empty."})
                continue
                
            if r_name not in resource_names:
                errors.append({"sheet": "Absences", "row": row_num, "column": "Resource Name", "message": f"Resource '{r_name}' is not defined in the Resources sheet."})
                
            parsed_date = None
            try:
                if isinstance(date_val, (datetime, pd.Timestamp)):
                    parsed_date = date_val.date()
                else:
                    parsed_date = datetime.strptime(str(date_val).strip(), "%Y-%m-%d").date()
            except Exception:
                errors.append({"sheet": "Absences", "row": row_num, "column": "Absence Date", "message": f"Invalid Date format '{date_val}'. Expected YYYY-MM-DD."})
                
            if parsed_date and start_date and end_date:
                if not (start_date <= parsed_date <= end_date):
                    errors.append({"sheet": "Absences", "row": row_num, "column": "Absence Date", "message": f"Absence date {parsed_date} is outside planning horizon ({start_date} to {end_date})."})
                    
            absences.append({
                "resource_name": r_name,
                "date": parsed_date,
                "reason": reason
            })
            
    if errors:
        raise ExcelValidationError(errors)
        
    return {
        "start_date": start_date,
        "end_date": end_date,
        "default_shift_hours": default_shift_hours,
        "resources": resources,
        "departments": departments,
        "mappings": mappings,
        "absences": absences
    }
